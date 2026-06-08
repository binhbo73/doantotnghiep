"""
Excel Chunker v2 - Row + Column Aware
=====================================

Creates one chunk per non-empty physical spreadsheet row and renders each row
with explicit Excel column letters. This keeps row/cell questions like "row 1"
or "A10" deterministic instead of depending on semantic vector search.
"""

import csv
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelChunkerV2:
    """Row + column aware Excel chunker."""

    def __init__(self):
        self.strategy_name = "excel_row_column_aware"

    def chunk_excel_file(
        self,
        file_path: str,
        document_id: str,
        metadata: Dict[str, Any] = None,
    ) -> List[Dict[str, Any]]:
        """
        Parse an Excel file and create one chunk per meaningful physical row.

        Args:
            file_path: Path to Excel file.
            document_id: Document ID.
            metadata: Additional metadata.

        Returns:
            List of row-aware chunks.
        """
        try:
            suffix = Path(file_path).suffix.lower()
            if suffix == ".csv":
                return self._chunk_csv_file(file_path, metadata)
            if suffix == ".xls":
                return self._chunk_xls_file(file_path, metadata)

            workbook = load_workbook(file_path, data_only=True)
            chunks: List[Dict[str, Any]] = []

            for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
                worksheet = workbook[sheet_name]
                chunks.extend(
                    self._chunk_worksheet(
                        worksheet=worksheet,
                        sheet_name=sheet_name,
                        sheet_idx=sheet_idx,
                        document_id=document_id,
                        metadata=metadata,
                    )
                )

            logger.info(
                "Excel Chunker v2: %s chunks from %s sheet(s)",
                len(chunks),
                len(workbook.sheetnames),
            )
            return chunks

        except Exception as e:
            logger.error("Excel chunking error: %s", str(e), exc_info=True)
            raise

    def _chunk_csv_file(
        self,
        file_path: str,
        metadata: Dict[str, Any] = None,
    ) -> List[Dict[str, Any]]:
        """Chunk CSV rows with the same metadata contract as Excel rows."""
        with open(file_path, "r", encoding="utf-8-sig", newline="") as csv_file:
            sample = csv_file.read(8192)
            csv_file.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            rows = list(csv.reader(csv_file, dialect))

        all_rows = []
        max_col = max((len(row) for row in rows), default=0)
        for row_idx, row in enumerate(rows, start=1):
            padded = list(row) + [""] * (max_col - len(row))
            if any(str(cell).strip() for cell in padded):
                all_rows.append({
                    "index": row_idx,
                    "data": padded,
                    "has_raw_value": True,
                    "inherited_columns": [],
                })

        return self._build_row_chunks(
            all_rows=all_rows,
            sheet_name=Path(file_path).stem or "CSV",
            sheet_idx=0,
            max_col=max_col,
            metadata=metadata,
            has_merged_cells=False,
        )

    def _chunk_xls_file(
        self,
        file_path: str,
        metadata: Dict[str, Any] = None,
    ) -> List[Dict[str, Any]]:
        """Chunk legacy XLS workbooks using xlrd."""
        import xlrd

        workbook = xlrd.open_workbook(file_path, formatting_info=True)
        chunks: List[Dict[str, Any]] = []
        for sheet_idx, sheet in enumerate(workbook.sheets()):
            merged_ranges = list(getattr(sheet, "merged_cells", []) or [])
            inherited_values = {}
            for row_low, row_high, col_low, col_high in merged_ranges:
                if row_high - row_low <= 1:
                    continue
                anchor = sheet.cell_value(row_low, col_low)
                for row_idx in range(row_low + 1, row_high):
                    inherited_values[(row_idx, col_low)] = anchor

            all_rows = []
            for zero_row_idx in range(sheet.nrows):
                row_data = []
                inherited_columns = []
                has_raw_value = False
                for col_idx in range(sheet.ncols):
                    value = sheet.cell_value(zero_row_idx, col_idx)
                    if value not in (None, ""):
                        has_raw_value = True
                    elif (zero_row_idx, col_idx) in inherited_values:
                        value = inherited_values[(zero_row_idx, col_idx)]
                        inherited_columns.append(get_column_letter(col_idx + 1))
                    row_data.append(
                        self._format_xls_value(
                            value,
                            sheet.cell_type(zero_row_idx, col_idx),
                            workbook.datemode,
                        )
                    )

                if any(str(cell).strip() for cell in row_data):
                    all_rows.append({
                        "index": zero_row_idx + 1,
                        "data": row_data,
                        "has_raw_value": has_raw_value,
                        "inherited_columns": inherited_columns,
                    })

            chunks.extend(self._build_row_chunks(
                all_rows=all_rows,
                sheet_name=sheet.name,
                sheet_idx=sheet_idx,
                max_col=sheet.ncols,
                metadata=metadata,
                has_merged_cells=bool(merged_ranges),
            ))

        logger.info(
            "Excel Chunker v2: %s chunks from %s legacy sheet(s)",
            len(chunks),
            workbook.nsheets,
        )
        return chunks

    def _format_xls_value(self, value: Any, cell_type: int, datemode: int) -> Any:
        """Preserve readable dates and integers from legacy XLS cells."""
        try:
            import xlrd

            if cell_type == xlrd.XL_CELL_DATE:
                return xlrd.xldate_as_datetime(value, datemode)
        except (TypeError, ValueError):
            pass
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    def _chunk_worksheet(
        self,
        worksheet,
        sheet_name: str,
        sheet_idx: int,
        document_id: str,
        metadata: Dict[str, Any] = None,
    ) -> List[Dict[str, Any]]:
        """
        Chunk a worksheet row by row.

        A row is included when at least one cell is non-empty. Row 1 is not
        treated specially because users usually mean the literal Excel row.
        """
        all_rows = []
        max_col = worksheet.max_column
        merged_ranges = list(worksheet.merged_cells.ranges)
        for row_idx in range(1, worksheet.max_row + 1):
            row_data = []
            inherited_columns = []
            has_raw_value = False

            for col_idx in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None and str(cell_value).strip():
                    has_raw_value = True

                if cell_value is None:
                    inherited_value = self._merged_row_value(
                        worksheet=worksheet,
                        merged_ranges=merged_ranges,
                        row_idx=row_idx,
                        col_idx=col_idx,
                    )
                    if inherited_value is not None:
                        cell_value = inherited_value
                        inherited_columns.append(get_column_letter(col_idx))

                row_data.append(cell_value if cell_value is not None else "")

            if row_data and any(str(cell).strip() for cell in row_data):
                all_rows.append(
                    {
                        "index": row_idx,
                        "data": row_data,
                        "has_raw_value": has_raw_value,
                        "inherited_columns": inherited_columns,
                    }
                )

        return self._build_row_chunks(
            all_rows=all_rows,
            sheet_name=sheet_name,
            sheet_idx=sheet_idx,
            max_col=max_col,
            metadata=metadata,
            has_merged_cells=bool(merged_ranges),
        )

    def _build_row_chunks(
        self,
        all_rows: List[Dict[str, Any]],
        sheet_name: str,
        sheet_idx: int,
        max_col: int,
        metadata: Dict[str, Any] = None,
        has_merged_cells: bool = False,
    ) -> List[Dict[str, Any]]:
        """Build whole-table and physical-row chunks from normalized rows."""
        chunks: List[Dict[str, Any]] = []
        if not all_rows or max_col <= 0:
            return chunks

        col_letters = [get_column_letter(i + 1) for i in range(max_col)]
        column_names = {
            col_idx: {"letter": col_letter, "name": col_letter}
            for col_idx, col_letter in enumerate(col_letters)
        }

        logger.info(
            "Sheet '%s': %s columns, %s non-empty rows",
            sheet_name,
            max_col,
            len(all_rows),
        )

        table_chunk = self._build_preserved_table_chunk(
            all_rows=all_rows,
            sheet_name=sheet_name,
            sheet_idx=sheet_idx,
            column_names=column_names,
            max_col=max_col,
            metadata=metadata,
            has_merged_cells=has_merged_cells,
        )
        if table_chunk:
            chunks.append(table_chunk)

        for data_row_idx, row_obj in enumerate(all_rows, start=0):
            row_idx = row_obj["index"]
            row_data = row_obj["data"]
            chunk_text = self._render_row_with_context(
                row_data=row_data,
                row_number=row_idx,
                sheet_name=sheet_name,
                column_names=column_names,
                max_col=max_col,
            )

            chunk_metadata = metadata.copy() if metadata else {}
            chunk_metadata.update(
                {
                    "sheet_name": sheet_name,
                    "sheet_idx": sheet_idx,
                    "row_number": row_idx,
                    "row_start": row_idx,
                    "row_end": row_idx,
                    "row_idx": data_row_idx,
                    "column_count": max_col,
                    "column_names": [
                        col_names.get("name", f"Col_{i}")
                        for i, col_names in column_names.items()
                    ],
                    "column_letters": col_letters,
                    "content_format": "spreadsheet_markdown",
                    "chunking_strategy": self.strategy_name,
                    "is_header_inclusive": False,
                    "has_merged_cells": has_merged_cells,
                    "is_merged_expanded_row": not row_obj.get("has_raw_value", True),
                    "merged_inherited_columns": row_obj.get("inherited_columns", []),
                }
            )

            chunks.append(
                {
                    "text": chunk_text,
                    "page_number": sheet_idx + 1,
                    "metadata": chunk_metadata,
                    "chunk_index": data_row_idx + (1 if table_chunk else 0),
                    "node_type": "detail",
                }
            )

        return chunks

    def _build_preserved_table_chunk(
        self,
        all_rows: List[Dict[str, Any]],
        sheet_name: str,
        sheet_idx: int,
        column_names: Dict[int, Dict[str, str]],
        max_col: int,
        metadata: Dict[str, Any] = None,
        has_merged_cells: bool = False,
    ) -> Optional[Dict[str, Any]]:
        """Create a whole-table chunk for small/medium spreadsheets."""
        col_letters = [
            column_names.get(col_idx, {}).get("letter", get_column_letter(col_idx + 1))
            for col_idx in range(max_col)
        ]
        lines = [f"# Sheet: {sheet_name}", ""]
        lines.append("| " + " | ".join(["Excel row"] + col_letters) + " |")
        lines.append("| " + " | ".join(["---"] * (max_col + 1)) + " |")
        for row_obj in all_rows:
            row_data = list(row_obj["data"])
            while len(row_data) < max_col:
                row_data.append("")
            row_values = [self._markdown_cell_value(value) for value in row_data[:max_col]]
            lines.append("| " + " | ".join([str(row_obj["index"])] + row_values) + " |")

        table_text = "\n".join(lines).strip()
        max_tokens = int(getattr(settings, "RAG_TABLE_CHUNK_MAX_TOKENS", 1600))
        if self._estimate_token_count(table_text) > max_tokens:
            return None

        chunk_metadata = metadata.copy() if metadata else {}
        chunk_metadata.update({
            "sheet_name": sheet_name,
            "sheet_idx": sheet_idx,
            "row_start": all_rows[0]["index"],
            "row_end": all_rows[-1]["index"],
            "row_count": len(all_rows),
            "column_count": max_col,
            "column_names": [
                col_names.get("name", f"Col_{i}")
                for i, col_names in column_names.items()
            ],
            "column_letters": col_letters,
            "content_format": "spreadsheet_markdown",
            "chunking_strategy": self.strategy_name,
            "table_preserved": True,
            "table_split": False,
            "table_scope": "worksheet",
            "is_header_inclusive": True,
            "has_merged_cells": has_merged_cells,
        })
        return {
            "text": table_text,
            "page_number": sheet_idx + 1,
            "metadata": chunk_metadata,
            "chunk_index": 0,
            "node_type": "detail",
        }

    def _merged_row_value(
        self,
        worksheet,
        merged_ranges: List[Any],
        row_idx: int,
        col_idx: int,
    ) -> Any:
        """Return the displayed merged value for row continuations.

        OpenPyXL stores a merged range value only in its top-left cell. For
        row-specific retrieval, continuation rows should still be searchable,
        but horizontal merged cells on the anchor row should not be duplicated
        across every covered column.
        """
        for merged_range in merged_ranges:
            if not (
                merged_range.min_row <= row_idx <= merged_range.max_row
                and merged_range.min_col <= col_idx <= merged_range.max_col
            ):
                continue
            if row_idx <= merged_range.min_row:
                return None
            if col_idx != merged_range.min_col:
                return None
            return worksheet.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
            ).value
        return None

    def _render_row_with_context(
        self,
        row_data: List[Any],
        row_number: int,
        sheet_name: str,
        column_names: Dict[int, Dict[str, str]],
        max_col: int,
    ) -> str:
        """
        Render a physical Excel row as Markdown.

        The first table uses the exact format expected by SpreadsheetRetriever:

        | Excel row | A | B | C |
        | --- | --- | --- | --- |
        | 9 | value A | value B | value C |
        """
        lines = [f"# Sheet: {sheet_name}, Row {row_number}", ""]

        padded_row = list(row_data)
        while len(padded_row) < max_col:
            padded_row.append("")

        col_letters = [
            column_names.get(col_idx, {}).get("letter", get_column_letter(col_idx + 1))
            for col_idx in range(max_col)
        ]
        row_values = [self._markdown_cell_value(value) for value in padded_row[:max_col]]

        lines.append("| " + " | ".join(["Excel row"] + col_letters) + " |")
        lines.append("| " + " | ".join(["---"] * (max_col + 1)) + " |")
        lines.append("| " + " | ".join([str(row_number)] + row_values) + " |")
        lines.append("")
        lines.append("| Column | Value |")
        lines.append("|---|---|")
        for col_letter, value in zip(col_letters, row_values):
            lines.append(f"| {col_letter} | {value} |")

        lines.append("")
        col_start = get_column_letter(1)
        col_end = get_column_letter(max_col)
        lines.append(f"**Position:** {col_start}{row_number}:{col_end}{row_number}")
        lines.append(f"**Context:** Row {row_number} of table on {sheet_name}")

        return "\n".join(lines)

    def _markdown_cell_value(self, value: Any) -> str:
        if isinstance(value, datetime):
            value = value.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(value, date):
            value = value.strftime("%Y-%m-%d")
        text = str(value).strip() if value is not None else ""
        if not text:
            return "(empty)"
        text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\n", " / ")
        return text.replace("|", "\\|")

    def _estimate_token_count(self, text: str) -> int:
        """Rough token estimation (1 token ~= 4 chars)."""
        return len(text) // 4


class ExcelTableExtractor:
    """
    Extract structured table data from Excel for RAG results.
    Useful when you want to return actual table structure, not markdown.
    """

    @staticmethod
    def extract_as_structured_table(
        file_path: str,
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Extract Excel as structured table data.

        Returns:
        {
            'sheets': [
                {
                    'name': 'Sheet2',
                    'headers': ['TT', 'Activity', 'Score'],
                    'rows': [
                        {'TT': '3', 'Activity': '...', 'Score': '30'},
                        ...
                    ]
                }
            ]
        }
        """
        workbook = load_workbook(file_path, data_only=True)
        sheets_data = []

        for sheet in workbook.sheetnames:
            if sheet_name and sheet_name != sheet:
                continue

            worksheet = workbook[sheet]
            rows = []
            headers = None

            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
                row_data = [cell if cell is not None else "" for cell in row]

                while row_data and not row_data[-1]:
                    row_data.pop()

                if row_idx == 0 and row_data:
                    headers = [str(h).strip() for h in row_data]
                elif row_data and any(row_data):
                    row_dict = {}
                    for i, header in enumerate(headers) if headers else []:
                        row_dict[header] = (
                            str(row_data[i]).strip() if i < len(row_data) else ""
                        )
                    rows.append(row_dict)

            sheets_data.append(
                {
                    "name": sheet,
                    "headers": headers or [],
                    "rows": rows,
                }
            )

        return {
            "file": file_path,
            "sheets": sheets_data,
        }
