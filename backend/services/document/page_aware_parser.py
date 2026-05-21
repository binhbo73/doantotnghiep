"""
Page-Aware Document Parser
===========================
Enhances standard parsing with explicit page boundary tracking.

Features:
- Extracts actual page numbers from PDF/DOCX
- Marks page boundaries in text
- Preserves paragraph structure
- Maps text spans to pages

For PDF: Uses PyPDF or pdf2image to extract page count + page breaks
For DOCX: Analyzes section/paragraph structure for logical pages
For Excel: Treats each sheet as a "page"
"""

import logging
import unicodedata
import re
from typing import List, Dict, Tuple, Optional, Any
from pathlib import Path
from datetime import date, datetime
from django.conf import settings

from services.document.office_preview import convert_office_to_pdf
from services.document.pdf_runtime import convert_pdf_to_markdown_quiet, read_pdf_page_counts, read_pdf_page_texts

logger = logging.getLogger(__name__)


class PageBoundary:
    """Represents a page boundary marker."""
    def __init__(self, page_number: int, char_start: int, char_end: int = None):
        self.page_number = page_number
        self.char_start = char_start
        self.char_end = char_end
    
    def __repr__(self):
        return f"Page{self.page_number}[{self.char_start}:{self.char_end}]"


class PageAwareText:
    """Text with explicit page mapping."""
    def __init__(self, text: str, boundaries: List[PageBoundary], total_pages: int, metadata: Optional[Dict[str, Any]] = None):
        self.text = text
        self.boundaries = sorted(boundaries, key=lambda b: b.char_start)
        self.total_pages = total_pages
        self.metadata = metadata or {}

        # Fill end positions so page lookup works reliably.
        for index, boundary in enumerate(self.boundaries):
            next_start = self.boundaries[index + 1].char_start if index + 1 < len(self.boundaries) else len(self.text)
            boundary.char_end = next_start
    
    def get_page_at_position(self, char_pos: int) -> int:
        """Get page number for a character position."""
        for boundary in self.boundaries:
            if boundary.char_start <= char_pos < (boundary.char_end or len(self.text)):
                return boundary.page_number
        return self.total_pages or 1
    
    def get_page_range(self, start_char: int, end_char: int) -> Tuple[int, int]:
        """Get start and end page numbers for a character range."""
        start_page = self.get_page_at_position(start_char)
        end_page = self.get_page_at_position(end_char - 1) if end_char > 0 else start_page
        return start_page, end_page


class PageAwareParserEnhancer:
    """
    Enhances existing parser output with page-aware information.
    
    Usage:
        enhancer = PageAwareParserEnhancer()
        raw_text, metadata = existing_parser.parse_file(path)
        page_aware = enhancer.enhance_pdf(raw_text, path)
        # Now use page_aware.get_page_at_position() for chunks
    """
    
    # PDF page marker (inserted by parse_pdf)
    PAGE_BREAK_MARKER = "\n\n--- [PAGE BREAK] ---\n\n"
    PAGE_BREAK_PATTERN = r"\n\n--- \[PAGE BREAK\] ---\n\n"
    
    def __init__(self):
        pass

    def _page_texts_to_page_aware(
        self,
        page_texts: List[str],
        source: str,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> Optional[PageAwareText]:
        """Build PageAwareText from already separated page text."""
        clean_pages = page_texts or []
        if not clean_pages or not any((page or '').strip() for page in clean_pages):
            return None

        text_parts = []
        boundaries = []
        char_pos = 0
        total_pages = len(clean_pages)
        for page_number, page_text in enumerate(clean_pages, start=1):
            page_header = f"\n--- Page {page_number} ---\n\n"
            text_parts.append(page_header)
            boundaries.append(PageBoundary(page_number, char_pos))
            char_pos += len(page_header)

            body = (page_text or '').strip()
            if body:
                body += "\n"
                text_parts.append(body)
                char_pos += len(body)

        return self._annotate_toc_metadata(PageAwareText(
            ''.join(text_parts),
            boundaries or [PageBoundary(1, 0)],
            max(1, total_pages),
            metadata={
                "page_count_source": source,
                **(metadata or {}),
            },
        ), clean_pages)

    def _normalize_toc_text(self, text: str) -> str:
        normalized = unicodedata.normalize('NFD', (text or '').lower())
        normalized = ''.join(ch for ch in normalized if unicodedata.category(ch) != 'Mn')
        return normalized.replace('đ', 'd')

    def _is_toc_like_page_text(self, text: str) -> bool:
        if not text or not text.strip():
            return False

        normalized = self._normalize_toc_text(text)
        if any(marker in normalized for marker in ('muc luc', 'table of contents', 'contents', 'index')):
            return True

        lines = [line.strip() for line in text.splitlines() if line.strip()]
        if len(lines) < 3:
            return False

        toc_like_lines = 0.0
        for line in lines[:30]:
            if re.search(r'\.{2,}\s*\d+\s*$', line) or re.search(r'\s\d+\s*$', line):
                toc_like_lines += 1.5
            elif len(line.split()) <= 10:
                toc_like_lines += 0.5

        ratio = toc_like_lines / max(1, min(len(lines), 30))
        return ratio >= 0.45

    def _detect_toc_pages(self, page_texts: List[str]) -> List[int]:
        toc_pages = []
        for index, page_text in enumerate(page_texts, start=1):
            if self._is_toc_like_page_text(page_text):
                toc_pages.append(index)
        return toc_pages

    def _annotate_toc_metadata(
        self,
        page_aware_text: PageAwareText,
        page_texts: Optional[List[str]] = None,
    ) -> PageAwareText:
        if not page_aware_text:
            return page_aware_text

        if page_texts is None:
            page_texts = [
                page_aware_text.text[boundary.char_start:boundary.char_end]
                for boundary in page_aware_text.boundaries
            ]

        toc_pages = self._detect_toc_pages(page_texts)
        metadata = dict(page_aware_text.metadata or {})
        if toc_pages:
            metadata['toc_pages'] = toc_pages
            metadata['has_toc'] = True
            if len(toc_pages) == len(page_aware_text.boundaries):
                metadata['layout_role'] = 'toc'
        page_aware_text.metadata = metadata
        return page_aware_text

    def _sanitize_pdf_markdown(self, text: str) -> str:
        """Remove embedded image payloads from markdown to keep chunking text-only."""
        if not text:
            return text

        original_len = len(text)
        sanitized = text

        # Replace markdown images with a short placeholder so alt text context is preserved.
        sanitized = re.sub(
            r'!\[([^\]]*)\]\(([^)]+)\)',
            lambda m: f"[Image: {(m.group(1) or 'embedded image').strip()}]",
            sanitized,
            flags=re.IGNORECASE,
        )

        # Replace HTML img tags with a stable placeholder.
        def _replace_img_tag(match):
            tag = match.group(0)
            alt_match = re.search(r'alt=[\"\']([^\"\']+)[\"\']', tag, flags=re.IGNORECASE)
            title_match = re.search(r'title=[\"\']([^\"\']+)[\"\']', tag, flags=re.IGNORECASE)
            label = None
            if alt_match and alt_match.group(1).strip():
                label = alt_match.group(1).strip()
            elif title_match and title_match.group(1).strip():
                label = title_match.group(1).strip()
            return f"[Image: {label or 'embedded image'}]"

        sanitized = re.sub(
            r'<img\b[^>]*>',
            _replace_img_tag,
            sanitized,
            flags=re.IGNORECASE,
        )

        # Safety net: remove any leftover data URI payload blocks.
        sanitized = re.sub(
            r'data:image\/[a-zA-Z0-9.+-]+;base64,[A-Za-z0-9+/=\s]+',
            '[EmbeddedImageData]',
            sanitized,
            flags=re.IGNORECASE,
        )

        sanitized = re.sub(r'\n{3,}', '\n\n', sanitized)
        sanitized = sanitized.strip()

        if len(sanitized) < original_len:
            logger.info(
                f"PDF markdown sanitized: {original_len} -> {len(sanitized)} chars "
                f"(removed {original_len - len(sanitized)} chars of image payloads)"
            )

        return sanitized
    
    def enhance_pdf(self, file_path: str) -> Optional[PageAwareText]:
        """
        Extract page boundaries from PDF.
        
        Returns:
            PageAwareText with page mapping
        """
        try:
            if getattr(settings, 'RAG_PDF_EXACT_PAGE_TEXT', True):
                page_aware = self._page_texts_to_page_aware(
                    read_pdf_page_texts(file_path),
                    source="pypdf_page_text",
                )
                if page_aware:
                    return page_aware

            import tempfile
            import opendataloader_pdf
            
            # Get page count first
            total_pages, page_char_counts = read_pdf_page_counts(file_path)
            
            # Parse with opendataloader-pdf
            with tempfile.TemporaryDirectory() as temp_dir:
                convert_pdf_to_markdown_quiet(
                    opendataloader_pdf,
                    input_path=[file_path],
                    output_dir=temp_dir,
                    format="markdown-with-images",
                    image_output="embedded",
                    image_format="png"
                )
                
                base_name = Path(file_path).stem
                md_file = Path(temp_dir) / f"{base_name}.md"
                
                if not md_file.exists():
                    logger.warning(f"No markdown output for {file_path}")
                    return None
                
                with open(md_file, 'r', encoding='utf-8') as f:
                    text = f.read()

            text = self._sanitize_pdf_markdown(text)
            
            # Insert page markers after parsing
            # Since opendataloader-pdf doesn't preserve page info in markdown,
            # we'll estimate based on text length per page
            # ✅ P0#1: Dùng tỉ lệ per-page text length từ pypdf (proportional)
            text_with_markers = self._insert_page_markers_proportional(
                text, page_char_counts
            )
            boundaries = self._extract_boundaries(text_with_markers)
            
            return self._annotate_toc_metadata(PageAwareText(text_with_markers, boundaries, total_pages))
        
        except Exception as e:
            logger.error(f"Error in PDF page-aware parsing: {str(e)}")
            return None

    def enhance_office_pdf(self, file_path: str) -> Optional[PageAwareText]:
        """Extract Office documents by converting to the same PDF preview used by the UI."""
        try:
            preview_pdf = convert_office_to_pdf(file_path)
            page_aware = self._page_texts_to_page_aware(
                read_pdf_page_texts(preview_pdf),
                source="office_pdf_preview",
                metadata={"preview_pdf_path": preview_pdf},
            )
            if page_aware:
                return self._annotate_toc_metadata(page_aware)
        except Exception as e:
            logger.warning(f"Office PDF page-aware parsing failed for {file_path}: {e}")
        return None
    
    def enhance_docx(self, file_path: str) -> Optional[PageAwareText]:
        """
        Extract page boundaries from DOCX with improved detection.
        """
        try:
            from docx import Document as DocxDoc
            
            docx = DocxDoc(file_path)
            
            text_parts = []
            page_num = 1
            boundaries = [PageBoundary(1, 0)] # Start with page 1
            char_pos = 0
            
            for para in docx.paragraphs:
                # 1. Check for page break in this paragraph
                if self._has_page_break(para):
                    page_num += 1
                    text_parts.append(self.PAGE_BREAK_MARKER)
                    char_pos += len(self.PAGE_BREAK_MARKER)
                    boundaries.append(PageBoundary(page_num, char_pos))
                
                para_text = para.text + "\n"
                text_parts.append(para_text)
                char_pos += len(para_text)
            
            # Handle tables
            for table in docx.tables:
                for row in table.rows:
                    for cell in row.cells:
                        text_parts.append(cell.text + " ")
                        char_pos += len(cell.text) + 1
                    text_parts.append("\n")
                    char_pos += 1
            
            full_text = "".join(text_parts)

            aligned = self._align_docx_text_to_pdf_preview(full_text, file_path)
            if aligned:
                return aligned
            
            # 2. FALLBACK: If document is long but no page breaks were found
            # Word pages are roughly 3000 characters. 
            if page_num == 1 and len(full_text) > 3500:
                logger.info(f"DOCX has no explicit breaks but is long ({len(full_text)} chars). Using logical page estimation.")
                return self._estimate_logical_pages(full_text)
            
            return PageAwareText(full_text, boundaries, page_num)
        
        except Exception as e:
            logger.error(f"Error in DOCX page-aware parsing: {str(e)}")
            return None

    def _align_docx_text_to_pdf_preview(self, docx_text: str, file_path: str) -> Optional[PageAwareText]:
        """Align DOCX extraction text to the same PDF preview pages shown in UI."""
        try:
            preview_pdf = convert_office_to_pdf(file_path)
            pdf_pages = read_pdf_page_texts(preview_pdf)
            pdf_pages = [page for page in pdf_pages if page and page.strip()]
            if not pdf_pages:
                return None

            docx_tokens = self._tokenize_with_positions(docx_text)
            pdf_page_tokens = [self._tokenize_plain(page_text) for page_text in pdf_pages]
            total_pdf_tokens = sum(len(tokens) for tokens in pdf_page_tokens)
            if not docx_tokens or total_pdf_tokens == 0:
                return None

            boundaries = [PageBoundary(1, 0)]
            current_token_index = 0
            cumulative_pdf_tokens = 0

            for page_index, page_tokens in enumerate(pdf_page_tokens[:-1], start=1):
                cumulative_pdf_tokens += len(page_tokens)
                estimated_index = int((cumulative_pdf_tokens / total_pdf_tokens) * len(docx_tokens))
                next_page_tokens = pdf_page_tokens[page_index]
                boundary_token_index = self._find_next_page_anchor(
                    [token for token, _ in docx_tokens],
                    next_page_tokens,
                    estimated_index,
                    current_token_index,
                )
                current_token_index = max(current_token_index + 1, boundary_token_index)
                boundaries.append(PageBoundary(page_index + 1, docx_tokens[current_token_index][1]))

            return self._annotate_toc_metadata(PageAwareText(
                docx_text,
                boundaries,
                len(pdf_pages),
                metadata={
                    "page_count_source": "office_pdf_preview_alignment",
                    "preview_pdf_path": preview_pdf,
                },
            ), pdf_pages)
        except Exception as e:
            logger.warning(f"Could not align DOCX pages to PDF preview: {e}")
            return None

    def _find_next_page_anchor(
        self,
        doc_tokens: List[str],
        next_page_tokens: List[str],
        estimated_index: int,
        min_index: int,
    ) -> int:
        anchor = [token for token in next_page_tokens if len(token) >= 2][:10]
        if len(anchor) < 4:
            return max(min_index, min(estimated_index, len(doc_tokens) - 1))

        search_radius = 500
        start = max(min_index, estimated_index - search_radius)
        end = min(len(doc_tokens) - len(anchor), estimated_index + search_radius)
        best_index = None
        best_score = 0

        for idx in range(start, max(start, end) + 1):
            window = doc_tokens[idx:idx + len(anchor)]
            score = sum(1 for left, right in zip(window, anchor) if left == right)
            if score > best_score:
                best_score = score
                best_index = idx
                if score == len(anchor):
                    break

        if best_index is not None and best_score >= max(4, len(anchor) // 2):
            return best_index

        return max(min_index, min(estimated_index, len(doc_tokens) - 1))

    def _tokenize_with_positions(self, text: str) -> List[Tuple[str, int]]:
        return [
            (self._normalize_token(match.group(0)), match.start())
            for match in re.finditer(r"\w+", text, flags=re.UNICODE)
            if self._normalize_token(match.group(0))
        ]

    def _tokenize_plain(self, text: str) -> List[str]:
        return [
            self._normalize_token(match.group(0))
            for match in re.finditer(r"\w+", text or "", flags=re.UNICODE)
            if self._normalize_token(match.group(0))
        ]

    def _normalize_token(self, token: str) -> str:
        normalized = unicodedata.normalize('NFD', token.lower())
        normalized = ''.join(ch for ch in normalized if unicodedata.category(ch) != 'Mn')
        return normalized.replace('đ', 'd')

    def _estimate_logical_pages(self, text: str, chars_per_page: int = 3000) -> PageAwareText:
        """Fallback to estimate pages based on character count for documents without markers."""
        total_pages = max(1, (len(text) + chars_per_page - 1) // chars_per_page)
        return self._annotate_toc_metadata(self._estimate_pages_by_total(text, total_pages, source="char_estimate"))

    def _estimate_pages_by_total(
        self,
        text: str,
        total_pages: int,
        source: str = "estimated",
    ) -> PageAwareText:
        """Create clean boundaries for a known total page count."""
        total_pages = max(1, int(total_pages or 1))
        chars_per_page = max(1, (len(text) + total_pages - 1) // total_pages)
        boundaries = []
        
        # Try to find paragraph breaks near each page interval.
        current_pos = 0
        for p in range(1, total_pages + 1):
            boundaries.append(PageBoundary(p, current_pos))
            
            target_next = p * chars_per_page
            if target_next >= len(text):
                break
            
            # Find nearest newline to make a clean break
            next_break = text.find('\n', target_next)
            if next_break != -1 and next_break - target_next < 500:
                current_pos = next_break + 1
            else:
                current_pos = target_next

        return self._annotate_toc_metadata(PageAwareText(text, boundaries, total_pages, metadata={"page_count_source": source}))
    
    def enhance_excel(self, file_path: str) -> Optional[PageAwareText]:
        """
        Extract page boundaries from Excel.
        
        Each sheet is treated as a page.
        """
        try:
            if Path(file_path).suffix.lower() == '.xls':
                return self._enhance_xls(file_path)

            import openpyxl
            from openpyxl.utils import get_column_letter
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            text_parts = []
            page_num = 1
            boundaries = []
            char_pos = 0
            sheet_stats = []
            total_non_empty_rows = 0
            total_non_empty_cells = 0
            total_merged_ranges = 0
            total_embedded_images = 0
            
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                
                # Add sheet header
                sheet_header = f"\n--- Sheet: {sheet} (Page {page_num}) ---\n\n"
                text_parts.append(sheet_header)
                boundaries.append(PageBoundary(page_num, char_pos))
                char_pos += len(sheet_header)

                min_row, max_row, min_col, max_col = self._excel_used_bounds(ws)
                if min_row is None:
                    sheet_stats.append({
                        "sheet_name": sheet,
                        "page_number": page_num,
                        "non_empty_rows": 0,
                        "non_empty_cells": 0,
                        "max_row": 0,
                        "max_col": 0,
                        "merged_ranges": 0,
                        "embedded_images": 0,
                    })
                    empty_text = "(empty sheet)\n\n"
                    text_parts.append(empty_text)
                    char_pos += len(empty_text)
                    page_num += 1
                    continue

                merge_ranges = [str(merge_range) for merge_range in ws.merged_cells.ranges]
                total_merged_ranges += len(merge_ranges)
                if merge_ranges:
                    merge_text = f"Merged cells: {', '.join(merge_ranges)}\n\n"
                    text_parts.append(merge_text)
                    char_pos += len(merge_text)

                image_cells = []
                for image in getattr(ws, "_images", []) or []:
                    marker = getattr(getattr(image, "anchor", None), "_from", None)
                    if marker is not None:
                        image_cells.append(f"{get_column_letter(marker.col + 1)}{marker.row + 1}")
                total_embedded_images += len(image_cells)
                if image_cells:
                    image_text = f"Embedded images anchored at: {', '.join(image_cells)}\n\n"
                    text_parts.append(image_text)
                    char_pos += len(image_text)

                col_letters = [get_column_letter(col_idx) for col_idx in range(min_col, max_col + 1)]
                header_cells = ["Excel row"] + col_letters
                separator_cells = ["---"] * len(header_cells)
                header_text = "| " + " | ".join(header_cells) + " |\n"
                separator_text = "| " + " | ".join(separator_cells) + " |\n"
                text_parts.append(header_text)
                text_parts.append(separator_text)
                char_pos += len(header_text) + len(separator_text)

                for row_idx in range(min_row, max_row + 1):
                    row_values = [str(row_idx)]
                    has_content = False
                    row_non_empty_cells = 0
                    for col_idx in range(min_col, max_col + 1):
                        value = self._format_excel_cell(ws.cell(row=row_idx, column=col_idx).value)
                        if value:
                            has_content = True
                            row_non_empty_cells += 1
                        row_values.append(value)

                    if not has_content:
                        continue
                    total_non_empty_rows += 1
                    total_non_empty_cells += row_non_empty_cells

                    row_text = "| " + " | ".join(row_values) + " |\n"
                    text_parts.append(row_text)
                    char_pos += len(row_text)

                text_parts.append("\n")
                char_pos += 1
                sheet_stats.append({
                    "sheet_name": sheet,
                    "page_number": page_num,
                    "non_empty_rows": sum(
                        1
                        for row_idx in range(min_row, max_row + 1)
                        if any(ws.cell(row=row_idx, column=col_idx).value is not None for col_idx in range(min_col, max_col + 1))
                    ),
                    "non_empty_cells": sum(
                        1
                        for row_idx in range(min_row, max_row + 1)
                        for col_idx in range(min_col, max_col + 1)
                        if ws.cell(row=row_idx, column=col_idx).value is not None
                    ),
                    "max_row": max_row,
                    "max_col": max_col,
                    "merged_ranges": len(merge_ranges),
                    "embedded_images": len(image_cells),
                })
                
                page_num += 1
            
            text = "".join(text_parts)
            
            spreadsheet_metadata = {
                "content_format": "spreadsheet_markdown",
                "sheet_count": len(workbook.sheetnames),
                "total_non_empty_rows": total_non_empty_rows,
                "total_non_empty_cells": total_non_empty_cells,
                "total_merged_ranges": total_merged_ranges,
                "total_embedded_images": total_embedded_images,
                "sheets": sheet_stats,
            }

            return PageAwareText(text, boundaries, page_num - 1, metadata=spreadsheet_metadata)
        
        except Exception as e:
            logger.error(f"Error in Excel page-aware parsing: {str(e)}")
            return None

    def _enhance_xls(self, file_path: str) -> Optional[PageAwareText]:
        """Extract text and sheet boundaries from legacy XLS workbooks."""
        try:
            import xlrd

            workbook = xlrd.open_workbook(file_path, formatting_info=False)
            text_parts = []
            boundaries = []
            char_pos = 0
            sheet_stats = []
            total_non_empty_rows = 0
            total_non_empty_cells = 0

            for sheet_index, sheet in enumerate(workbook.sheets(), start=1):
                sheet_header = f"\n--- Sheet: {sheet.name} (Page {sheet_index}) ---\n\n"
                text_parts.append(sheet_header)
                boundaries.append(PageBoundary(sheet_index, char_pos))
                char_pos += len(sheet_header)

                col_letters = [self._excel_column_letter(col_idx + 1) for col_idx in range(sheet.ncols)]
                header_cells = ["Excel row"] + col_letters
                separator_cells = ["---"] * len(header_cells)
                header_text = "| " + " | ".join(header_cells) + " |\n"
                separator_text = "| " + " | ".join(separator_cells) + " |\n"
                text_parts.append(header_text)
                text_parts.append(separator_text)
                char_pos += len(header_text) + len(separator_text)

                non_empty_rows = 0
                non_empty_cells = 0
                for row_idx in range(sheet.nrows):
                    row_values = []
                    has_content = False
                    for col_idx in range(sheet.ncols):
                        value = self._format_excel_cell(sheet.cell_value(row_idx, col_idx))
                        if value:
                            has_content = True
                            non_empty_cells += 1
                        row_values.append(value)

                    if not has_content:
                        continue

                    non_empty_rows += 1
                    row_text = "| " + " | ".join([str(row_idx + 1)] + row_values) + " |\n"
                    text_parts.append(row_text)
                    char_pos += len(row_text)

                text_parts.append("\n")
                char_pos += 1
                total_non_empty_rows += non_empty_rows
                total_non_empty_cells += non_empty_cells
                sheet_stats.append({
                    "sheet_name": sheet.name,
                    "page_number": sheet_index,
                    "non_empty_rows": non_empty_rows,
                    "non_empty_cells": non_empty_cells,
                    "max_row": sheet.nrows,
                    "max_col": sheet.ncols,
                    "merged_ranges": 0,
                    "embedded_images": 0,
                })

            text = "".join(text_parts)
            total_pages = max(1, workbook.nsheets)
            return PageAwareText(
                text,
                boundaries or [PageBoundary(1, 0)],
                total_pages,
                metadata={
                    "content_format": "spreadsheet_markdown",
                    "sheet_count": workbook.nsheets,
                    "total_non_empty_rows": total_non_empty_rows,
                    "total_non_empty_cells": total_non_empty_cells,
                    "total_merged_ranges": 0,
                    "total_embedded_images": 0,
                    "sheets": sheet_stats,
                },
            )
        except Exception as e:
            logger.error(f"Error in XLS page-aware parsing: {str(e)}")
            return None

    def _excel_column_letter(self, index: int) -> str:
        """Convert a 1-based Excel column index to a column letter."""
        letters = []
        while index > 0:
            index, remainder = divmod(index - 1, 26)
            letters.append(chr(65 + remainder))
        return ''.join(reversed(letters)) or 'A'

    def _excel_used_bounds(self, ws):
        """Return used worksheet bounds including merged cells and image anchors."""
        min_row = min_col = None
        max_row = max_col = None

        def include_cell(row_idx: int, col_idx: int):
            nonlocal min_row, min_col, max_row, max_col
            min_row = row_idx if min_row is None else min(min_row, row_idx)
            max_row = row_idx if max_row is None else max(max_row, row_idx)
            min_col = col_idx if min_col is None else min(min_col, col_idx)
            max_col = col_idx if max_col is None else max(max_col, col_idx)

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    include_cell(cell.row, cell.column)

        for merge_range in ws.merged_cells.ranges:
            for row_idx in (merge_range.min_row, merge_range.max_row):
                for col_idx in (merge_range.min_col, merge_range.max_col):
                    include_cell(row_idx, col_idx)

        for image in getattr(ws, "_images", []) or []:
            marker = getattr(getattr(image, "anchor", None), "_from", None)
            if marker is not None:
                include_cell(marker.row + 1, marker.col + 1)

        return min_row, max_row, min_col, max_col

    def _format_excel_cell(self, value: Any) -> str:
        """Format an Excel cell value for markdown-table chunks."""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        text = str(value).replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
        text = re.sub(r"\s+", " ", text).strip()
        return text.replace("|", "\\|")
    
    def enhance_text(self, file_path: str) -> Optional[PageAwareText]:
        """
        For plain text: estimate pages based on content length.
        Assume ~2000 chars per page (typical).
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()
            
            # Estimate pages
            chars_per_page = 2000
            total_pages = max(1, (len(text) + chars_per_page - 1) // chars_per_page)
            
            # Create boundaries
            boundaries = []
            for page in range(1, total_pages + 1):
                start_char = (page - 1) * chars_per_page
                boundaries.append(PageBoundary(page, start_char))
            
            return PageAwareText(text, boundaries, total_pages)
        
        except Exception as e:
            logger.error(f"Error in text page-aware parsing: {str(e)}")
            return None

    def enhance_csv(self, file_path: str, rows_per_page: int = 100) -> Optional[PageAwareText]:
        """Treat CSV as paged row groups so the chunker can keep page-like locality."""
        try:
            import csv

            text_parts = []
            boundaries = []
            char_pos = 0
            page_num = 1

            with open(file_path, 'r', encoding='utf-8-sig', newline='') as csv_file:
                reader = csv.reader(csv_file)
                for row_index, row in enumerate(reader):
                    if row_index % rows_per_page == 0:
                        boundaries.append(PageBoundary(page_num, char_pos))
                        text_parts.append(f"\n--- CSV Page {page_num} ---\n")
                        char_pos += len(text_parts[-1])
                        page_num += 1

                    row_text = " | ".join(cell.strip() for cell in row) + "\n"
                    text_parts.append(row_text)
                    char_pos += len(row_text)

            text = "".join(text_parts)
            total_pages = max(1, page_num - 1)
            return PageAwareText(text, boundaries, total_pages)
        except Exception as e:
            logger.error(f"Error in CSV page-aware parsing: {str(e)}")
            return None
    
    def _insert_page_markers(self, text: str, total_pages: int) -> str:
        """
        Heuristically insert page markers into parsed text.
        
        Strategy: Estimate page boundaries based on text length
        """
        if total_pages <= 1:
            return text
        
        chars_per_page = len(text) // total_pages
        markers = []
        
        for page in range(1, total_pages):
            # Find good break point (paragraph, sentence, or exact position)
            target_pos = page * chars_per_page
            break_pos = self._find_good_break_point(text, target_pos)
            markers.append((break_pos, page))
        
        # Insert markers in reverse order to maintain positions
        result = text
        for pos, page in sorted(markers, reverse=True):
            result = result[:pos] + self.PAGE_BREAK_MARKER + result[pos:]
        
        return result
    
    def _find_good_break_point(self, text: str, target_pos: int, window: int = 200) -> int:
        """Find a good place to break (paragraph or sentence)."""
        start = max(0, target_pos - window)
        end = min(len(text), target_pos + window)
        
        segment = text[start:end]
        
        # Look for paragraph break first (double newline)
        para_break = segment.rfind('\n\n')
        if para_break != -1:
            return start + para_break
        
        # Look for sentence end (period followed by space)
        sentence_ends = [m.end() for m in re.finditer(r'\.\s', segment)]
        if sentence_ends:
            closest = min(sentence_ends, key=lambda x: abs(x - (target_pos - start)))
            return start + closest
        
        # Fallback to target position
        return target_pos
    
    # ✅ P0#1: NEW method — proportional page marker placement
    def _insert_page_markers_proportional(
        self, text: str, page_char_counts: list
    ) -> str:
        """
        Insert page markers using per-page text length ratios from pypdf.
        
        Khác với _insert_page_markers (uniform division), phương thức này dùng tỉ lệ
        text thực tế mỗi trang từ pypdf để phân bổ vị trí marker trong Markdown output.
        
        Ví dụ: PDF 3 trang, pypdf trả về char counts [5000, 2000, 3000]
        → Tổng = 10000, tỉ lệ 50%/20%/30%
        → Markdown dài 20000 chars → markers tại 10000, 14000 (theo tỉ lệ)
        
        Args:
            text: Full Markdown text từ opendataloader-pdf
            page_char_counts: List[int] char count mỗi trang từ pypdf
        
        Returns:
            Text with proportional page break markers inserted
        """
        total_pages = len(page_char_counts)
        if total_pages <= 1:
            return text
        
        total_chars = sum(page_char_counts)
        if total_chars == 0:
            # Fallback to uniform if all pages are empty (edge case)
            return self._insert_page_markers_legacy(text, total_pages)
        
        text_len = len(text)
        markers = []
        cumulative_chars = 0
        
        for page_idx in range(total_pages - 1):
            cumulative_chars += page_char_counts[page_idx]
            ratio = cumulative_chars / total_chars
            target_pos = int(ratio * text_len)
            # Clamp to valid range
            target_pos = max(1, min(text_len - 1, target_pos))
            break_pos = self._find_good_break_point(text, target_pos)
            markers.append((break_pos, page_idx + 1))
        
        # Insert markers in reverse order to maintain positions
        result = text
        for pos, page in sorted(markers, reverse=True):
            result = result[:pos] + self.PAGE_BREAK_MARKER + result[pos:]
        
        logger.debug(
            f"Proportional markers: {total_pages} pages, "
            f"ratios={[round(c/total_chars, 2) for c in page_char_counts]}"
        )
        return result

    def _insert_page_markers_legacy(self, text: str, total_pages: int) -> str:
        """Fallback: uniform division (giữ lại để backward compatibility)."""
        return self._insert_page_markers(text, total_pages)

    def _extract_boundaries(self, text: str) -> List[PageBoundary]:
        """Extract page boundaries from text with markers."""
        boundaries = [PageBoundary(1, 0)]
        page_num = 2

        # Each page-break marker separates previous page from next page.
        # So the next page starts AFTER the marker token.
        for match in re.finditer(self.PAGE_BREAK_PATTERN, text):
            next_page_start = min(len(text), match.end())
            # Skip duplicate/non-increasing starts to keep boundaries valid.
            if next_page_start > boundaries[-1].char_start:
                boundaries.append(PageBoundary(page_num, next_page_start))
                page_num += 1

        return boundaries
    
    def _has_page_break(self, paragraph) -> bool:
        """
        Check if paragraph has a page break (manual or rendered).
        
        Detection logic:
        1. Explicit pageBreakBefore in paragraph properties
        2. <w:br w:type="page"/> in any run
        3. <w:lastRenderedPageBreak/> in any run (inserted by Word when paginating)
        """
        try:
            # 1. Check paragraph properties for "Page Break Before"
            if paragraph._element.pPr is not None:
                pPr = paragraph._element.pPr
                # Looking for <w:pageBreakBefore/>
                if pPr.find('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}pageBreakBefore') is not None:
                    return True

            # 2. Check each run for break elements
            for run in paragraph.runs:
                # Check for <w:br w:type="page"/> (Manual page break)
                if 'lastRenderedPageBreak' in run._element.xml or 'w:br' in run._element.xml:
                    if run._element.find('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}lastRenderedPageBreak') is not None:
                        return True
                    
                    brs = run._element.findall('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}br')
                    for br in brs:
                        if br.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}type') == 'page':
                            return True
            return False
        except Exception:
            return False
