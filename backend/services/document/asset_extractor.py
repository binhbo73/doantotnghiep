"""
Asset Extractor - Trích xuất ảnh từ PDF, DOCX, XLSX
====================================================

Trả về danh sách dict thông tin asset:
{
    'image_data': bytes,
    'image_format': str,
    'asset_type': str,
    'page_number': int | None,
    'sheet_name': str | None,
    'anchor_cell': str | None,
    'anchor_row': int | None,
    'paragraph_index': int | None,
    'position': dict,
    'context_text': str,
    'width': int,
    'height': int,
    'source': str,
}
"""

import io
import os
import zipfile
import re
import logging
from typing import List, Dict, Any, Optional
from pathlib import Path

logger = logging.getLogger(__name__)


class AssetExtractor:
    """Trích xuất ảnh từ các định dạng tài liệu."""

    IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.tif', '.webp', '.emf', '.wmf'}
    CONTEXT_CHARS_BEFORE = 500
    CONTEXT_CHARS_AFTER = 200

    def __init__(self, media_root: str = None):
        from django.conf import settings
        self.media_root = media_root or os.path.join(
            getattr(settings, 'MEDIA_ROOT', 'media'), 'document_assets'
        )

    # ═══════════════════════════════════════════════════════════════
    # PUBLIC API
    # ═══════════════════════════════════════════════════════════════

    def extract_from_pdf(self, file_path: str, full_text: str = '') -> List[Dict[str, Any]]:
        """
        Trích xuất ảnh từ PDF bằng PyMuPDF.

        Hai chiến lược:
        1. PDF có embedded images → extract từng ảnh
        2. PDF scan / không có embedded → render từng trang thành ảnh
        """
        assets = []
        try:
            import fitz  # PyMuPDF
            from django.conf import settings

            doc = fitz.open(file_path)
            total_pages = len(doc)
            render_text_pages = getattr(settings, 'ASSET_RENDER_TEXT_PDF_PAGES', False)
            scanned_text_threshold = int(getattr(settings, 'ASSET_SCANNED_PAGE_TEXT_THRESHOLD', 40))

            for page_num in range(total_pages):
                page = doc[page_num]

                # ── Chiến lược 1: Extract embedded images ──
                image_list = page.get_images(full=True)

                if image_list:
                    for img_index, img_info in enumerate(image_list):
                        try:
                            xref = img_info[0]
                            base_image = doc.extract_image(xref)
                            image_bytes = base_image["image"]
                            image_ext = base_image["ext"]

                            rects = page.get_image_rects(xref)
                            position = {
                                'x': float(rects[0][0]) if rects else 0,
                                'y': float(rects[0][1]) if rects else 0,
                                'width': float(rects[0][2]) if rects else 0,
                                'height': float(rects[0][3]) if rects else 0,
                            }

                            context_text = self._get_pdf_context(full_text, page_num + 1, position)

                            assets.append({
                                'image_data': image_bytes,
                                'image_format': image_ext,
                                'asset_type': 'pdf_embedded',
                                'page_number': page_num + 1,
                                'position': position,
                                'context_text': context_text,
                                'width': base_image.get('width', 0),
                                'height': base_image.get('height', 0),
                                'source': f'page_{page_num + 1}_img_{img_index}',
                            })
                        except Exception as e:
                            logger.warning(
                                f"PDF embedded image error (page {page_num + 1}, img {img_index}): {e}"
                            )
                            continue

                # ── Chiến lược 2: Render toàn trang ──
                # Render scanned pages only. Rendering every text page without
                # embedded raster images turns an 11-page PDF into 11 assets.
                page_text = (page.get_text("text") or "").strip()
                should_render_page = (
                    not image_list
                    and (render_text_pages or len(page_text) <= scanned_text_threshold)
                )

                if should_render_page:
                    try:
                        pix = page.get_pixmap(dpi=150)
                        image_bytes = pix.tobytes("png")

                        assets.append({
                            'image_data': image_bytes,
                            'image_format': 'png',
                            'asset_type': 'pdf_page_render',
                            'page_number': page_num + 1,
                            'position': {
                                'x': 0, 'y': 0,
                                'width': float(pix.width),
                                'height': float(pix.height),
                            },
                            'context_text': self._get_page_context(full_text, page_num + 1),
                            'width': pix.width,
                            'height': pix.height,
                            'source': f'page_{page_num + 1}_render',
                        })
                    except Exception as e:
                        logger.warning(f"PDF page render error (page {page_num + 1}): {e}")
                        continue

            doc.close()
            logger.info(f"PDF extraction: {len(assets)} assets from {total_pages} pages")

        except ImportError:
            logger.error("PyMuPDF (fitz) not installed. Install: pip install PyMuPDF")
        except Exception as e:
            logger.error(f"PDF asset extraction failed: {e}", exc_info=True)

        return assets

    def extract_from_docx(self, file_path: str, full_text: str = '') -> List[Dict[str, Any]]:
        """
        Trích xuất ảnh từ DOCX.

        DOCX là file ZIP: word/document.xml (nội dung) + word/media/* (ảnh).
        """
        assets = []
        try:
            # ── Đọc quan hệ ảnh từ python-docx ──
            from docx import Document

            doc = Document(file_path)
            para_texts = [p.text for p in doc.paragraphs]

            # Map rId → image blob
            image_map = {}
            for rel in doc.part.rels.values():
                if "image" in rel.reltype:
                    image_map[rel.rId] = {
                        'blob': rel.target_part.blob,
                        'content_type': rel.target_part.content_type or 'image/png',
                    }

            if not image_map:
                # Fallback: đọc trực tiếp word/media/ từ ZIP
                image_map = self._fallback_docx_zip(file_path)

            if not image_map:
                logger.info(f"No images found in DOCX: {file_path}")
                return assets

            # ── Đọc document.xml tìm vị trí ảnh ──
            with zipfile.ZipFile(file_path, 'r') as z:
                document_xml = z.read('word/document.xml').decode('utf-8')

            # Tìm <wp:inline> hoặc <wp:anchor> → <a:blip r:embed="rIdX">
            paragraph_blocks = re.findall(r'<w:p\b.*?</w:p>', document_xml, re.DOTALL)
            img_entries = []
            for xml_para_idx, paragraph_xml in enumerate(paragraph_blocks):
                for rId in re.findall(r'<a:blip[^>]*(?:r:embed|r:link)="(rId\d+)"', paragraph_xml):
                    img_entries.append((xml_para_idx, rId))

            if not img_entries:
                img_pattern = re.compile(
                    r'<(?:wp:inline|wp:anchor).*?'
                    r'<a:blip[^>]*(?:r:embed|r:link)="(rId\d+)"[^>]*>'
                    r'.*?</(?:wp:inline|wp:anchor)>',
                    re.DOTALL,
                )
                img_refs = img_pattern.findall(document_xml)
                img_entries = [
                    (min(idx, len(para_texts) - 1) if para_texts else 0, rId)
                    for idx, rId in enumerate(img_refs)
                ]

            for idx, (para_idx, rId) in enumerate(img_entries):
                if rId not in image_map:
                    continue

                img_info = image_map[rId]
                image_bytes = img_info['blob']

                # Xác định format
                fmt = self._guess_format(img_info.get('content_type', ''))

                # Xác định paragraph gần nhất
                para_idx = min(para_idx, len(para_texts) - 1) if para_texts else 0

                # Context text
                context_text = ''
                if 0 <= para_idx < len(para_texts):
                    before = para_texts[para_idx - 1][-200:] if para_idx > 0 else ''
                    current = para_texts[para_idx]
                    after = para_texts[para_idx + 1][:200] if para_idx + 1 < len(para_texts) else ''
                    context_text = f"{before}\n{current}\n{after}"

                # Kích thước ảnh
                width, height = self._get_image_size(image_bytes)

                assets.append({
                    'image_data': image_bytes,
                    'image_format': fmt,
                    'asset_type': 'docx_inline',
                    'paragraph_index': para_idx,
                    'context_text': context_text,
                    'position': {
                        'paragraph_index': para_idx,
                        'image_index': idx,
                        'width': width,
                        'height': height,
                    },
                    'width': width,
                    'height': height,
                    'source': f'paragraph_{para_idx}_rId_{rId}',
                })

            logger.info(f"DOCX extraction: {len(assets)} assets from {file_path}")

        except Exception as e:
            logger.error(f"DOCX asset extraction failed: {e}", exc_info=True)

        return assets

    def extract_from_xlsx(self, file_path: str, full_text: str = '') -> List[Dict[str, Any]]:
        """
        Trích xuất ảnh từ XLSX.

        XLSX là file ZIP: xl/worksheets/sheet*.xml + xl/drawings/drawing*.xml + xl/media/*.
        """
        assets = []
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter

            wb = openpyxl.load_workbook(file_path, data_only=True)

            # ── Cách 1: Dùng openpyxl _images ──
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                for image in getattr(ws, '_images', []) or []:
                    try:
                        image_bytes = image._data()
                        fmt = getattr(image, 'format', 'png').lower()

                        anchor = getattr(image, 'anchor', None)
                        anchor_cell = None
                        anchor_row = None
                        if anchor:
                            from_cell = getattr(anchor, '_from', None)
                            if from_cell:
                                # Logic: Excel offsets are in EMUs (12700 EMUs = 1 point)
                                # We shift the anchor if offsets exceed cell dimensions
                                col = from_cell.col
                                row = from_cell.row
                                
                                col_off = getattr(from_cell, 'colOff', 0) or 0
                                row_off = getattr(from_cell, 'rowOff', 0) or 0
                                
                                # Shift Column if colOff is large (multi-cell shift)
                                if col_off > 0:
                                    curr_col = col
                                    while col_off > 0:
                                        try:
                                            col_width_chars = ws.column_dimensions[get_column_letter(curr_col + 1)].width or 8.43
                                            # Approx column width in points: width * 7 pixels, pixels / 1.33 = points.
                                            # So width * 5.25 approx points.
                                            col_width_emus = int(col_width_chars * 5.25 * 12700)
                                            if col_off >= col_width_emus and col_width_emus > 0:
                                                col_off -= col_width_emus
                                                curr_col += 1
                                            else:
                                                break
                                        except:
                                            break
                                    col = curr_col

                                # Shift Row if rowOff is large (multi-cell shift)
                                if row_off > 0:
                                    curr_row = row
                                    while row_off > 0:
                                        try:
                                            # row height is in points, convert to EMUs
                                            row_height_pts = ws.row_dimensions[curr_row + 1].height or 15
                                            row_height_emus = int(row_height_pts * 12700)
                                            if row_off >= row_height_emus and row_height_emus > 0:
                                                row_off -= row_height_emus
                                                curr_row += 1
                                            else:
                                                break
                                        except:
                                            break
                                    row = curr_row
                                    
                                col_letter = get_column_letter(col + 1)
                                anchor_cell = f"{col_letter}{row + 1}"
                                anchor_row = row + 1

                        context_text = self._get_xlsx_context(full_text, sheet_name, anchor_row)
                        width, height = self._get_image_size(image_bytes)

                        assets.append({
                            'image_data': image_bytes,
                            'image_format': fmt,
                            'asset_type': 'xlsx_image',
                            'sheet_name': sheet_name,
                            'anchor_cell': anchor_cell,
                            'anchor_row': anchor_row,
                            'context_text': context_text,
                            'position': {},
                            'width': width,
                            'height': height,
                            'source': f'sheet_{sheet_name}_cell_{anchor_cell}',
                        })
                    except Exception as e:
                        logger.warning(f"XLSX image error (sheet {sheet_name}): {e}")
                        continue

            # ── Cách 2: Đọc drawing XML ──
            if not assets:
                drawing_assets = self._fallback_xlsx_drawings(file_path, wb, full_text)
                assets.extend(drawing_assets)

            logger.info(f"XLSX extraction: {len(assets)} assets from {file_path}")

        except Exception as e:
            logger.error(f"XLSX asset extraction failed: {e}", exc_info=True)

        return assets

    # ═══════════════════════════════════════════════════════════════
    # SAVE
    # ═══════════════════════════════════════════════════════════════

    def save_asset_image(
        self, document_id: str, asset_data: Dict[str, Any], asset_index: int
    ) -> str:
        """Lưu ảnh vào media/document_assets/{doc_id}/... Trả về đường dẫn tương đối."""
        doc_dir = os.path.join(self.media_root, str(document_id))
        os.makedirs(doc_dir, exist_ok=True)

        ext = asset_data.get('image_format', 'png')
        source = asset_data.get('source', f'asset_{asset_index}')
        safe_source = "".join(c for c in source if c.isalnum() or c in '._-')
        filename = f"{asset_index:04d}_{safe_source}.{ext}"
        filepath = os.path.join(doc_dir, filename)

        with open(filepath, 'wb') as f:
            f.write(asset_data['image_data'])

        relative_path = os.path.join('document_assets', str(document_id), filename)
        logger.debug(f"Saved asset image: {filepath} ({len(asset_data['image_data'])} bytes)")
        return relative_path

    # ═══════════════════════════════════════════════════════════════
    # CONTEXT HELPERS
    # ═══════════════════════════════════════════════════════════════

    def _get_pdf_context(self, full_text: str, page_num: int, position: dict) -> str:
        if not full_text:
            return ''
        page_marker = "--- [PAGE BREAK] ---"
        pages = full_text.split(page_marker)
        if page_num <= len(pages):
            page_text = pages[page_num - 1]
            y_ratio = position.get('y', 0) / max(position.get('height', 1), 1)
            char_pos = int(len(page_text) * y_ratio)
            start = max(0, char_pos - self.CONTEXT_CHARS_BEFORE)
            end = min(len(page_text), char_pos + self.CONTEXT_CHARS_AFTER)
            return page_text[start:end]
        return full_text[:self.CONTEXT_CHARS_BEFORE]

    def _get_page_context(self, full_text: str, page_num: int) -> str:
        if not full_text:
            return ''
        page_marker = "--- [PAGE BREAK] ---"
        pages = full_text.split(page_marker)
        if page_num <= len(pages):
            return pages[page_num - 1][:1000]
        return full_text[:1000]

    def _get_xlsx_context(self, full_text: str, sheet_name: str, anchor_row: int) -> str:
        if not full_text or not sheet_name:
            return full_text[:500] if full_text else ''

        sheet_pattern = re.compile(
            rf"--- Sheet:\s*{re.escape(sheet_name)}\s*\(Page\s*\d+\)\s*---"
        )
        match = sheet_pattern.search(full_text)
        if not match:
            return full_text[:500]

        sheet_start = match.end()
        next_sheet = re.search(r"--- Sheet:", full_text[sheet_start:])
        sheet_end = sheet_start + next_sheet.start() if next_sheet else len(full_text)
        sheet_text = full_text[sheet_start:sheet_end]

        if anchor_row:
            lines = sheet_text.split('\n')
            for line in lines:
                row_match = re.match(rf'^\|\s*{anchor_row}\s*\|', line)
                if row_match:
                    idx = lines.index(line)
                    context_lines = lines[max(0, idx - 2):idx + 5]
                    return '\n'.join(context_lines)

        return sheet_text[:500]

    # ═══════════════════════════════════════════════════════════════
    # FALLBACKS
    # ═══════════════════════════════════════════════════════════════

    def _fallback_docx_zip(self, file_path: str) -> Dict[str, Dict]:
        """Fallback: đọc ảnh trực tiếp từ word/media/ trong ZIP."""
        image_map = {}
        try:
            with zipfile.ZipFile(file_path, 'r') as z:
                media_files = [
                    name for name in z.namelist()
                    if name.startswith('word/media/')
                    and os.path.splitext(name)[1].lower() in self.IMAGE_EXTENSIONS
                ]
                for idx, name in enumerate(media_files):
                    image_map[f"fallback_{idx}"] = {
                        'blob': z.read(name),
                        'content_type': f'image/{os.path.splitext(name)[1].lstrip(".")}',
                    }
        except Exception as e:
            logger.warning(f"DOCX ZIP fallback error: {e}")
        return image_map

    def _fallback_xlsx_drawings(
        self, file_path: str, wb, full_text: str
    ) -> List[Dict[str, Any]]:
        """Fallback: đọc drawing XML để map ảnh trong XLSX."""
        import xml.etree.ElementTree as ET
        from openpyxl.utils import get_column_letter

        assets = []
        try:
            with zipfile.ZipFile(file_path, 'r') as z:
                all_files = z.namelist()

                media_files = {}
                for name in all_files:
                    if name.startswith('xl/media/') and os.path.splitext(name)[1].lower() in self.IMAGE_EXTENSIONS:
                        media_files[os.path.basename(name)] = z.read(name)

                # Map sheet drawing relationships correctly
                # In XLSX, worksheets/_rels/sheetN.xml.rels maps rId to drawings/drawingM.xml
                sheet_drawing_map = {}
                for name in all_files:
                    if name.startswith('xl/worksheets/_rels/sheet') and name.endswith('.xml.rels'):
                        try:
                            sheet_idx_match = re.search(r'sheet(\d+)', name)
                            if not sheet_idx_match: continue
                            sheet_idx = int(sheet_idx_match.group(1)) - 1
                            if sheet_idx >= len(wb.sheetnames): continue
                            
                            rel_xml = z.read(name).decode('utf-8')
                            rel_root = ET.fromstring(rel_xml)
                            for rel in rel_root:
                                target = rel.get('Target', '')
                                if 'drawings/drawing' in target:
                                    drawing_name = os.path.basename(target)
                                    sheet_drawing_map[drawing_name] = wb.sheetnames[sheet_idx]
                        except Exception as e:
                            logger.warning(f"Error mapping sheet to drawing: {e}")

                drawing_files = [f for f in all_files if f.startswith('xl/drawings/drawing') and f.endswith('.xml')]

                for drawing_path in drawing_files:
                    try:
                        drawing_xml = z.read(drawing_path).decode('utf-8')
                        root = ET.fromstring(drawing_xml)

                        ns = {
                            'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                            'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                            'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
                        }

                        for anchor in (
                            root.findall('.//xdr:twoCellAnchor', ns) +
                            root.findall('.//xdr:oneCellAnchor', ns)
                        ):
                            from_elem = anchor.find('.//xdr:from', ns)
                            if from_elem is None:
                                continue

                            col_elem = from_elem.find('.//xdr:col', ns)
                            row_elem = from_elem.find('.//xdr:row', ns)
                            if col_elem is None or row_elem is None:
                                continue

                            col = int(col_elem.text)
                            row = int(row_elem.text)

                            # Handle offsets in fallback as well
                            col_off_elem = from_elem.find('.//xdr:colOff', ns)
                            row_off_elem = from_elem.find('.//xdr:rowOff', ns)
                            col_off = int(col_off_elem.text) if col_off_elem is not None else 0
                            row_off = int(row_off_elem.text) if row_off_elem is not None else 0

                            if col_off > 0:
                                curr_col = col
                                while col_off > 0:
                                    try:
                                        # Fallback to first sheet's dimensions
                                        col_width_chars = wb[wb.sheetnames[0]].column_dimensions[get_column_letter(curr_col + 1)].width or 8.43
                                        col_width_emus = int(col_width_chars * 5.25 * 12700)
                                        if col_off >= col_width_emus and col_width_emus > 0:
                                            col_off -= col_width_emus
                                            curr_col += 1
                                        else: break
                                    except: break
                                col = curr_col

                            if row_off > 0:
                                curr_row = row
                                while row_off > 0:
                                    try:
                                        row_height_pts = wb[wb.sheetnames[0]].row_dimensions[curr_row + 1].height or 15
                                        row_height_emus = int(row_height_pts * 12700)
                                        if row_off >= row_height_emus and row_height_emus > 0:
                                            row_off -= row_height_emus
                                            curr_row += 1
                                        else: break
                                    except: break
                                row = curr_row

                            cell_ref = f"{get_column_letter(col + 1)}{row + 1}"

                            blip = anchor.find('.//a:blip', ns)
                            if blip is None:
                                continue

                            embed_id = blip.get(f'{{{ns["r"]}}}embed')
                            if not embed_id:
                                continue

                            # Map sheet using the correctly parsed relationship map
                            drawing_filename = os.path.basename(drawing_path)
                            sheet_name = sheet_drawing_map.get(drawing_filename)
                            
                            if not sheet_name:
                                # Fallback to index if mapping failed
                                sheet_num = re.search(r'drawing(\d+)', drawing_path)
                                sheet_idx = int(sheet_num.group(1)) - 1 if sheet_num else 0
                                sheet_name = wb.sheetnames[sheet_idx] if sheet_idx < len(wb.sheetnames) else None

                            # Đọc rels
                            rels_path = drawing_path.replace(
                                'xl/drawings/', 'xl/drawings/_rels/'
                            ).replace('.xml', '.xml.rels')
                            if rels_path not in all_files:
                                continue

                            rels_xml = z.read(rels_path).decode('utf-8')
                            rels_root = ET.fromstring(rels_xml)
                            for rel in rels_root:
                                if rel.get('Id') == embed_id:
                                    target = rel.get('Target')
                                    media_name = os.path.basename(target)
                                    if media_name in media_files:
                                        image_bytes = media_files[media_name]
                                        ext = os.path.splitext(media_name)[1].lstrip('.').lower()
                                        width, height = self._get_image_size(image_bytes)
                                        context_text = self._get_xlsx_context(
                                            full_text, sheet_name, row + 1
                                        )

                                        assets.append({
                                            'image_data': image_bytes,
                                            'image_format': ext,
                                            'asset_type': 'xlsx_image',
                                            'sheet_name': sheet_name,
                                            'anchor_cell': cell_ref,
                                            'anchor_row': row + 1,
                                            'context_text': context_text,
                                            'position': {},
                                            'width': width,
                                            'height': height,
                                            'source': f'drawing_{media_name}',
                                        })
                    except Exception as e:
                        logger.warning(f"XLSX drawing XML parse error: {e}")
                        continue

        except Exception as e:
            logger.warning(f"XLSX drawing fallback error: {e}")

        return assets

    # ═══════════════════════════════════════════════════════════════
    # UTILITIES
    # ═══════════════════════════════════════════════════════════════

    @staticmethod
    def _guess_format(content_type: str) -> str:
        if 'jpeg' in content_type or 'jpg' in content_type:
            return 'jpg'
        if 'png' in content_type:
            return 'png'
        if 'gif' in content_type:
            return 'gif'
        if 'bmp' in content_type:
            return 'bmp'
        if 'webp' in content_type:
            return 'webp'
        return 'png'

    @staticmethod
    def _get_image_size(image_bytes: bytes) -> tuple:
        try:
            from PIL import Image
            img = Image.open(io.BytesIO(image_bytes))
            return img.size
        except Exception:
            return 0, 0
